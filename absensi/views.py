from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.db import OperationalError, ProgrammingError
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.utils.timezone import make_aware, is_naive
from datetime import datetime, timedelta
from xml.sax.saxutils import escape
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import json
import re
import pytesseract
from PIL import Image
import cv2
import numpy as np
from io import BytesIO
import os
from django.conf import settings
from .models import AbsensiTamu
from .forms import AbsensiForm, DateRangeForm, AdminProfileForm, AddAdminForm
from .ocr_utils import preprocess_and_ocr

# Konfigurasi path Tesseract untuk Windows
# Jika Tesseract sudah di PATH, tidak perlu set ini
# Jika belum, uncomment dan sesuaikan path sesuai instalasi Tesseract Anda
# Contoh path umum di Windows:
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
# Atau jika di Laragon:
# pytesseract.pytesseract.tesseract_cmd = r'C:\laragon\bin\tesseract\tesseract.exe'

# Coba auto-detect Tesseract path untuk Windows
if os.name == 'nt':  # Windows
    possible_paths = [
        r'C:\Program Files\Tesseract-OCR\tesseract.exe',
        r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
        r'C:\laragon\bin\tesseract\tesseract.exe',
    ]
    # Tambahkan path untuk user-specific jika USERNAME tersedia
    username = os.getenv('USERNAME') or os.getenv('USER')
    if username:
        possible_paths.append(
            r'C:\Users\{}\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'.format(username)
        )
    
    for path in possible_paths:
        if os.path.exists(path):
            pytesseract.pytesseract.tesseract_cmd = path
            break


def custom_404(request, exception=None):
    """Custom 404 error page - catch all unmatched URLs"""
    return render(request, 'absensi/404.html', status=404)


def test_404(request):
    """Route untuk testing halaman 404"""
    return render(request, 'absensi/404.html', status=404)


def index(request):
    """Halaman utama - hanya form input untuk non-user"""
    # Jika form di-submit
    if request.method == 'POST':
        form = AbsensiForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Data absensi berhasil disimpan!')
                return redirect('index')
            except (OperationalError, ProgrammingError) as e:
                messages.error(request, 'Database error! Silakan jalankan migrasi terlebih dahulu.')
    else:
        form = AbsensiForm()
    
    return render(request, 'absensi/index.html', {
        'form': form,
    })


def is_likely_id_number(text):
    """
    Menentukan apakah teks kemungkinan adalah nomor ID atau bukan.
    Nomor ID biasanya memiliki banyak angka (lebih dari 8 digit berurutan).
    """
    # Cari pola angka panjang (lebih dari 8 digit berurutan)
    long_number_pattern = r'\d{9,}'
    if re.search(long_number_pattern, text):
        return True
    
    # Cek apakah teks memiliki proporsi angka yang sangat tinggi (> 50%)
    digit_count = len(re.findall(r'\d', text))
    total_chars = len(re.findall(r'\w', text))
    if total_chars > 0 and digit_count / total_chars > 0.5 and digit_count >= 10:
        return True
    
    # Cek pola seperti "WIP. 199002122016121001" atau "ID: 1234567890"
    id_patterns = [
        r'\b(wip|id|nip|nik|no\.?|nrp)[\s:\.]+\d{8,}',
        r'[A-Z]{1,5}[\s\.]+\d{10,}',
    ]
    for pattern in id_patterns:
        if re.search(pattern, text, re.IGNORECASE):
            return True
    
    return False


def is_likely_person_name(text):
    """
    Menentukan apakah teks kemungkinan adalah nama orang.
    Nama orang biasanya:
    - Hanya huruf (boleh ada titik, koma, tanda hubung)
    - Tidak terlalu panjang (> 50 karakter)
    - Tidak terlalu pendek (< 3 karakter)
    - Tidak mengandung banyak angka
    - Tidak seperti ID atau nomor
    """
    # Exclude jika ini kemungkinan ID
    if is_likely_id_number(text):
        return False
    
    # Exclude jika terlalu panjang (biasanya nama tidak lebih dari 50 karakter tanpa spasi)
    if len(text) > 50:
        return False
    
    # Exclude jika terlalu pendek
    if len(text.strip()) < 3:
        return False
    
    # Exclude jika mengandung terlalu banyak angka (lebih dari 20% angka)
    digit_count = len(re.findall(r'\d', text))
    letter_count = len(re.findall(r'[a-zA-Z]', text))
    total_chars = letter_count + digit_count
    if total_chars > 0 and digit_count / total_chars > 0.2:
        return False
    
    # Exclude jika mengandung angka panjang (> 6 digit)
    if re.search(r'\d{7,}', text):
        return False
    
    # Harus minimal mengandung huruf (nama pasti ada hurufnya)
    if letter_count < 2:
        return False
    
    # Kata kunci yang menandakan ini bukan nama orang
    not_name_keywords = [
        'id', 'nip', 'nik', 'nrp', 'no.', 'tanggal', 'tanggal', 'tempat',
        'lahir', 'alamat', 'wip', 'card', 'kartu', 'pegawai', 'no:', 'no ',
        'kode', 'code', 'nomor', 'number'
    ]
    text_lower = text.lower()
    for keyword in not_name_keywords:
        if keyword in text_lower:
            return False
    
    return True


def detect_name_vs_institution(text_lines):
    """
    Mendeteksi mana yang nama orang dan mana yang nama instansi/perusahaan.
    Menggunakan berbagai heuristik untuk menentukan.
    """
    # Kata-kata kunci yang menandakan instansi/perusahaan
    institution_keywords = [
        'pt', 'cv', 'tbk', 'persero', 'kementerian', 'dinas', 'badan', 'lembaga',
        'departemen', 'kantor', 'perusahaan', 'instansi', 'organisasi', 'yayasan',
        'foundation', 'corp', 'inc', 'ltd', 'company', 'co', 'pemerintah',
        'pemkot', 'pemkab', 'pemprov', 'dpr', 'mpr', 'dpd', 'kpu', 'bawaslu',
        'pemerintah', 'negara', 'republik', 'provinsi', 'kabupaten', 'kota'
    ]
    
    # Kata-kata yang menandakan ini adalah jabatan (bukan nama instansi tapi terkait)
    position_keywords = [
        'direktur', 'manager', 'kepala', 'ketua', 'sekretaris', 'staff', 'karyawan',
        'pegawai', 'anggota', 'presiden', 'gubernur', 'bupati', 'walikota'
    ]
    
    cleaned_lines = []
    for line in text_lines:
        # Bersihkan baris dari karakter aneh tapi tetap pertahankan huruf dan angka penting
        clean_line = re.sub(r'[^\w\s\-\.\,]', '', line.strip())
        if clean_line and len(clean_line) > 2:
            cleaned_lines.append(clean_line)
    
    if not cleaned_lines:
        return None, None
    
    # Filter baris yang jelas bukan nama (ID numbers, dll)
    candidate_names = []
    candidate_institutions = []
    
    for idx, line in enumerate(cleaned_lines):
        line_lower = line.lower()
        
        # Skip jika ini jelas nomor ID
        if is_likely_id_number(line):
            continue
        
        # Cek apakah ini kemungkinan instansi
        has_inst_keyword = any(keyword in line_lower for keyword in institution_keywords)
        word_count = len(line.split())
        
        if has_inst_keyword or word_count > 3:
            # Ini kemungkinan instansi
            candidate_institutions.append({'line': line, 'idx': idx, 'score': 0})
        elif is_likely_person_name(line):
            # Ini kemungkinan nama orang
            candidate_names.append({'line': line, 'idx': idx, 'score': 0})
    
    # Hitung skor untuk kandidat nama
    for candidate in candidate_names:
        score = 0
        line = candidate['line']
        line_lower = line.lower()
        
        # Skor berdasarkan posisi (nama biasanya di atas)
        if candidate['idx'] < len(cleaned_lines) / 2:
            score += 2
        
        # Skor berdasarkan panjang (nama biasanya 3-40 karakter)
        if 3 <= len(line) <= 40:
            score += 1
        
        # Skor jika tidak ada keyword instansi
        if not any(keyword in line_lower for keyword in institution_keywords):
            score += 1
        
        # Skor jika hanya huruf (minimal angka)
        if len(re.findall(r'[a-zA-Z]', line)) / max(len(line), 1) > 0.8:
            score += 1
        
        # Skor berdasarkan jumlah kata (nama biasanya 2-4 kata)
        word_count = len(line.split())
        if 2 <= word_count <= 4:
            score += 1
        
        candidate['score'] = score
    
    # Hitung skor untuk kandidat instansi
    for candidate in candidate_institutions:
        score = 0
        line = candidate['line']
        line_lower = line.lower()
        
        # Skor berdasarkan keyword instansi
        for keyword in institution_keywords:
            if keyword in line_lower:
                score += 2
        
        # Skor berdasarkan posisi (instansi biasanya di bawah)
        if candidate['idx'] >= len(cleaned_lines) / 2:
            score += 1
        
        # Skor berdasarkan panjang (instansi bisa lebih panjang)
        if len(line) > 5:
            score += 0.5
        
        candidate['score'] = score
    
    # Ambil nama dengan skor tertinggi
    name_line = None
    if candidate_names:
        sorted_names = sorted(candidate_names, key=lambda x: x['score'], reverse=True)
        name_line = sorted_names[0]['line']
    
    # Gabungkan instansi jika ada beberapa baris yang relevan
    institution_line = None
    if candidate_institutions:
        sorted_inst = sorted(candidate_institutions, key=lambda x: x['score'], reverse=True)
        
        # Jika ada beberapa baris instansi dengan skor tinggi, gabungkan
        high_score_inst = [inst for inst in sorted_inst if inst['score'] > 0]
        if len(high_score_inst) > 1:
            # Gabungkan baris yang relevan (ambil hingga 4 baris teratas dengan skor > 0)
            inst_lines = []
            for inst in sorted_inst:
                if inst['score'] > 0 and len(inst_lines) < 4:
                    # Jangan tambahkan jika terlalu pendek atau sama dengan baris lain
                    if len(inst['line']) > 3 and inst['line'] not in inst_lines:
                        inst_lines.append(inst['line'])
            if inst_lines:
                institution_line = ' '.join(inst_lines)
        elif len(sorted_inst) > 0:
            # Jika hanya ada satu atau beberapa dengan skor rendah, gabungkan juga
            inst_lines = []
            for inst in sorted_inst[:3]:  # Ambil 3 teratas
                if inst['line'] not in inst_lines and len(inst['line']) > 3:
                    inst_lines.append(inst['line'])
            if inst_lines:
                if len(inst_lines) > 1:
                    institution_line = ' '.join(inst_lines)
                else:
                    institution_line = inst_lines[0]
    
    # Fallback: jika tidak ada yang terdeteksi dengan baik
    if not name_line and cleaned_lines:
        # Cari baris pertama yang bukan ID dan tidak terlalu panjang
        for line in cleaned_lines:
            if not is_likely_id_number(line) and is_likely_person_name(line):
                name_line = line
                break
    
    if not institution_line and cleaned_lines:
        # Cari baris yang bukan nama dan bukan ID
        for line in cleaned_lines:
            if line != name_line and not is_likely_id_number(line):
                if len(line.split()) > 2 or any(kw in line.lower() for kw in institution_keywords):
                    institution_line = line
                    break
    
    return name_line, institution_line


def analyze_with_gemini(ocr_text):
    """
    TAHAP 2: Menggunakan Gemini AI untuk menganalisis teks OCR dan mengekstrak nama dan instansi.
    
    Alur:
    1. Terima semua informasi OCR (mentah, acak)
    2. Analisis dengan Gemini AI
    3. Sortir dan pisahkan nama vs instansi
    4. Return JSON dengan nama dan instansi yang sudah dipisahkan
    """
    try:
        import google.generativeai as genai
        
        # Cek apakah API key tersedia
        api_key = getattr(settings, 'GEMINI_API_KEY', '')
        if not api_key:
            return None, None, "GEMINI_API_KEY tidak dikonfigurasi. Silakan set API key di settings.py atau environment variable."
        
        # Konfigurasi Gemini
        genai.configure(api_key=api_key)
        
        # List model Gemini yang akan dicoba (dari terbaru ke legacy)
        # Menggunakan model Gemini 2.5 dan 2.0 sesuai dokumentasi terbaru
        # Coba semua model sampai ada yang berfungsi!
        models_to_try = [
            # Gemini 2.5 Series (Terbaru - 2025)
            'gemini-2.5-pro',              # State-of-the-art, paling canggih
            'gemini-2.5-flash',            # Fast and intelligent, best price-performance
            'gemini-2.5-flash-lite',       # Ultra fast, cost-efficient
        ]
        
        model = None
        model_name = None
        last_error = None
        
        # Coba setiap model sampai ada yang berfungsi
        for model_name_attempt in models_to_try:
            try:
                print(f"[Gemini] Mencoba model: {model_name_attempt}")
                model = genai.GenerativeModel(model_name_attempt)
                model_name = model_name_attempt
                print(f"[Gemini] ✅ Model {model_name_attempt} berhasil!")
                break  # Berhasil, keluar dari loop
            except Exception as e:
                last_error = str(e)
                print(f"[Gemini] ❌ Model {model_name_attempt} gagal: {last_error}")
                continue  # Coba model berikutnya
        
        # Jika semua model gagal
        if model is None:
            return None, None, f"Semua model Gemini tidak tersedia. Model yang dicoba: {', '.join(models_to_try)}. Error terakhir: {last_error}"
        
        # Prompt untuk Gemini AI - Fleksibel untuk semua jenis gambar dengan teks
        prompt = f"""Anda adalah sistem AI yang ahli dalam menganalisis teks dari berbagai jenis dokumen/gambar yang sudah di-OCR.

TUGAS UTAMA:
Dari teks OCR berikut (yang bisa berasal dari ID card, kartu pegawai, kartu nama, dokumen, atau gambar apapun dengan teks), 
identifikasi dan ekstrak DUA informasi penting:
1. NAMA LENGKAP ORANG (nama manusia, bisa nama pegawai, nama pemilik, nama person, dll)
2. NAMA INSTANSI/PERUSAHAAN/ORGANISASI (nama institusi, perusahaan, organisasi, instansi, tempat kerja, dll)

TEKS OCR YANG PERLU DIANALISIS (semua teks yang berhasil di-OCR dari gambar):
{ocr_text}

CATATAN PENTING:
- Teks di atas bisa berasal dari berbagai jenis dokumen/gambar (ID card, kartu pegawai, kartu nama, dokumen resmi, surat, dll)
- Tidak masalah jika ada teks tambahan seperti alamat, nomor, tanggal, dll - Anda tetap harus ekstrak NAMA ORANG dan NAMA INSTANSI saja
- Fokus pada menemukan 2 informasi penting: NAMA ORANG dan NAMA INSTANSI/PERUSAHAAN
- Abaikan informasi lain seperti alamat, nomor telepon, email, tanggal, nomor ID, dll

PANDUAN IDENTIFIKASI NAMA ORANG (PENTING!):
✅ CIRI-CIRI NAMA ORANG (berlaku untuk semua jenis dokumen):
- Terdiri dari 2-5 KATA (misal: "Budi Santoso", "Ahmad Fauzi", "Siti Nurhaliza", "AFIF RAHMAN HAKIM", "Dr. John Smith", "Prof. Dr. Ahmad")
- Hanya huruf alfabet (A-Z, a-z), boleh titik (.), koma (,), tanda hubung (-), apostrof (')
- Panjangnya antara 3-60 karakter
- BISA mengandung gelar/title: "Dr.", "Prof.", "AMd", "S.E", "S.Kom", "M.M", "M.Sc", "Ph.D", dll
- TIDAK mengandung kata seperti: "Pemerintah", "Provinsi", "Kementerian", "Dinas", "PT", "CV", "Badan", "Lembaga", "Departemen", "Kantor", "Perusahaan", "Instansi", "Organisasi", "Jl.", "No.", "RT", "RW", "Kecamatan", "Kelurahan"
- Biasanya berupa nama manusia: bisa nama Indonesia (Budi, Ahmad, Siti) atau nama internasional (John, Smith, David)

❌ BUKAN NAMA ORANG:
- Instansi/perusahaan seperti "PEMERINTAH PROVINSI RIAU", "DINAS PERHUBUNGAN", "PT. Maju Jaya", "CV. Abadi"
- Nomor ID seperti "WIP. 199002122016121001", "ID: 1234567890", "NIP: 123456", "NIK: 123456789"
- Alamat seperti "Jl. Sudirman No. 123", "RT 05 RW 03", "Kecamatan X"
- Tanggal seperti "01 Januari 1990", "01-01-1990"
- Email, nomor telepon, kode pos
- Baris yang mengandung kata kunci instansi (Pemerintah, Dinas, PT, CV, Kementerian, dll)
- Jabatan saja tanpa nama seperti "Direktur", "Manager", "Kepala Dinas" (kecuali jika ada nama di sampingnya)

PANDUAN IDENTIFIKASI NAMA INSTANSI/PERUSAHAAN/ORGANISASI (PENTING!):
✅ CIRI-CIRI NAMA INSTANSI (berlaku untuk semua jenis dokumen):
- Biasanya lebih panjang (biasanya 2+ kata, bisa sampai 10+ kata)
- MENGANDUNG kata kunci instansi: PT, CV, TBK, Persero, Kementerian, Dinas, Badan, Lembaga, Departemen, Kantor, Perusahaan, Instansi, Organisasi, Yayasan, Foundation, Pemerintah, Pemkot, Pemkab, Pemprov, Provinsi, Kabupaten, Kota, Company, Corp, Inc, Ltd, Organization, Agency, Office, Department, Ministry
- Contoh: "PEMERINTAH PROVINSI RIAU", "DINAS PERHUBUNGAN", "PT. Maju Jaya Bersama", "Kementerian Komunikasi dan Informatika", "Google Inc.", "Microsoft Corporation"
- Bisa terdiri dari beberapa baris yang harus digabungkan menjadi satu teks lengkap
- TIDAK termasuk nama orang di akhir (jika ada nama orang di akhir baris instansi, PISAHKAN dan HAPUS dari instansi)
- Bisa berupa nama organisasi formal atau informal

❌ BUKAN NAMA INSTANSI:
- Nomor ID, NIP, NIK
- Nama orang saja (2-4 kata tanpa keyword instansi)
- Alamat lengkap
- Jabatan atau posisi kerja

CONTOH ANALISIS YANG BENAR (berbagai format dokumen):

Contoh 1 - ID Card:
OCR: "PEMERINTAH PROVINSI RIAU\nDINAS PERHUBUNGAN\nAFIF RAHMAN HAKIM, AMd. LLASOP\nWIP. 199002122016121001\nJl. Sudirman No. 123"
→ NAMA ORANG: "AFIF RAHMAN HAKIM, AMd. LLASOP" (atau "AFIF RAHMAN HAKIM")
→ INSTANSI: "PEMERINTAH PROVINSI RIAU DINAS PERHUBUNGAN"

Contoh 2 - Kartu Pegawai:
OCR: "Budi Santoso\nWIP. 199002122016121001\nPT. Maju Jaya Bersama\nJakarta"
→ NAMA ORANG: "Budi Santoso"
→ INSTANSI: "PT. Maju Jaya Bersama"

Contoh 3 - Dokumen Surat:
OCR: "Kepada Yth. Dr. Ahmad Fauzi\nDirektur PT. Teknologi Indonesia\nJl. Merdeka No. 45\nJakarta"
→ NAMA ORANG: "Dr. Ahmad Fauzi"
→ INSTANSI: "PT. Teknologi Indonesia"

Contoh 4 - Kartu Nama:
OCR: "John Smith\nSales Manager\nABC Company Inc.\njohn@abc.com\n+62 123 456 789"
→ NAMA ORANG: "John Smith"
→ INSTANSI: "ABC Company Inc."

Contoh 5 - Teks tercampur:
OCR: "PEMERINTAH PROVINSI RIAU DINAS PERHUBUNGAN AFIF RAHMAN HAKIM\nJl. Raya No. 1 RT 05 RW 03"
→ NAMA ORANG: "AFIF RAHMAN HAKIM" (extract dari tengah/akhir)
→ INSTANSI: "PEMERINTAH PROVINSI RIAU DINAS PERHUBUNGAN" (tanpa nama orang dan alamat)

INSTRUKSI KHUSUS YANG WAJIB DIIKUTI:
1. Analisis SEMUA teks yang ada, tidak peduli format dokumennya (ID card, kartu nama, surat, dll)
2. NAMA ORANG harus 2-5 kata, TIDAK boleh mengandung keyword instansi (Pemerintah, Dinas, PT, CV, dll)
3. INSTANSI harus mengandung keyword instansi atau berupa nama organisasi/perusahaan
4. Jika nama instansi dan nama orang digabung dalam satu baris, PISAHKAN:
   - Ambil bagian yang mengandung keyword instansi sebagai INSTANSI
   - Ambil bagian yang berupa nama orang (2-5 kata) sebagai NAMA ORANG
5. Abaikan informasi lain seperti: alamat, nomor telepon, email, tanggal, nomor ID, kode pos, dll
6. JANGAN masukkan nama orang ke dalam instansi
7. JANGAN masukkan instansi ke dalam nama orang
8. Jika ada beberapa nama orang, ambil yang paling lengkap/konsisten
9. Jika ada beberapa nama instansi, gabungkan menjadi satu jika saling terkait
10. Jika tidak yakin atau tidak ditemukan, gunakan string kosong "" (jangan mengarang)

BALAS HANYA DALAM FORMAT JSON (tanpa penjelasan, tanpa markdown):
{{
    "nama": "Nama Lengkap Pegawai (2-4 kata, tanpa instansi) atau string kosong",
    "instansi": "Nama Lengkap Instansi/Perusahaan (tanpa nama orang) atau string kosong"
}}

Contoh output yang BENAR untuk teks di atas:
{{
    "nama": "AFIF RAHMAN HAKIM",
    "instansi": "PEMERINTAH PROVINSI RIAU DINAS PERHUBUNGAN"
}}

JANGAN seperti ini (SALAH):
{{
    "nama": "PEMERINTAH PROVINSI RIAU",
    "instansi": "PEMERINTAH PROVINSI RIAU DINAS PERHUBUNGAN AFIF RAHMAN HAKIM"
}}

Sekarang analisis teks OCR berikut dan ekstrak dengan benar:"""
        
        # Panggil Gemini API dengan model yang berhasil
        try:
            print(f"[Gemini] Menggunakan model: {model_name}")
            response = model.generate_content(prompt)
            print(f"[Gemini] ✅ Response berhasil diterima dari model {model_name}")
        except Exception as e:
            # Jika model yang dipilih gagal saat generate, coba model lain
            print(f"[Gemini] ⚠️ Model {model_name} gagal saat generate: {str(e)}")
            print(f"[Gemini] Mencoba model alternatif...")
            
            # Coba model alternatif (prioritas model terbaru)
            alternative_models = [
                'gemini-2.5-flash',        # Terbaru
                'gemini-2.5-flash-lite',   # Terbaru lite
                'gemini-2.0-flash',        # Previous gen
                'gemini-1.5-flash',        # Stable
                'gemini-1.5-pro',          # Stable pro
                'gemini-pro',              # Legacy
            ]
            for alt_model_name in alternative_models:
                if alt_model_name == model_name:
                    continue  # Skip yang sudah dicoba
                try:
                    print(f"[Gemini] Mencoba model alternatif: {alt_model_name}")
                    alt_model = genai.GenerativeModel(alt_model_name)
                    response = alt_model.generate_content(prompt)
                    model_name = alt_model_name
                    print(f"[Gemini] ✅ Model alternatif {alt_model_name} berhasil!")
                    break
                except Exception as alt_e:
                    print(f"[Gemini] ❌ Model alternatif {alt_model_name} gagal: {str(alt_e)}")
                    continue
            else:
                # Semua model alternatif gagal
                raise Exception(f"Semua model gagal saat generate content. Error terakhir: {str(e)}")
        
        # Parse response
        response_text = response.text.strip()
        
        # Bersihkan response dari markdown code block jika ada
        if response_text.startswith('```json'):
            response_text = response_text[7:]
        elif response_text.startswith('```'):
            response_text = response_text[3:]
        if response_text.endswith('```'):
            response_text = response_text[:-3]
        response_text = response_text.strip()
        
        # Parse JSON
        detected_name = ''
        detected_institution = ''
        
        try:
            # Coba parse JSON langsung
            result = json.loads(response_text)
            detected_name = result.get('nama', '').strip()
            detected_institution = result.get('instansi', '').strip()
        except json.JSONDecodeError:
            # Jika JSON parsing gagal, coba extract manual
            # Cari JSON object dalam response
            json_start = response_text.find('{')
            json_end = response_text.rfind('}') + 1
            
            if json_start >= 0 and json_end > json_start:
                try:
                    json_text = response_text[json_start:json_end]
                    result = json.loads(json_text)
                    detected_name = result.get('nama', '').strip()
                    detected_institution = result.get('instansi', '').strip()
                except:
                    # Jika masih gagal, extract manual dari teks
                    lines = response_text.split('\n')
                    for line in lines:
                        line_lower = line.lower()
                        # Cari "nama" atau "name"
                        if ('nama' in line_lower or 'name' in line_lower) and ':' in line:
                            parts = line.split(':', 1)
                            if len(parts) == 2:
                                value = parts[1].strip().strip('"').strip("'").strip(',')
                                if value and not detected_name:
                                    detected_name = value
                        # Cari "instansi" atau "institution" atau "perusahaan" atau "company"
                        elif (('instansi' in line_lower or 'institution' in line_lower or 
                               'perusahaan' in line_lower or 'company' in line_lower) and ':' in line):
                            parts = line.split(':', 1)
                            if len(parts) == 2:
                                value = parts[1].strip().strip('"').strip("'").strip(',')
                                if value and not detected_institution:
                                    detected_institution = value
            
        # Validasi dan bersihkan hasil
        # Filter nama: harus minimal 3 karakter, tidak boleh nomor ID, TIDAK boleh mengandung keyword instansi
        institution_keywords_check = [
            'pemerintah', 'provinsi', 'kementerian', 'dinas', 'badan', 'lembaga',
            'departemen', 'kantor', 'perusahaan', 'instansi', 'organisasi',
            'yayasan', 'pemkot', 'pemkab', 'pemprov', 'kabupaten', 'kota',
            'pt', 'cv', 'tbk', 'persero'
        ]
        
        if detected_name:
            detected_name_lower = detected_name.lower()
            
            # Remove jika terlalu pendek
            if len(detected_name) < 3:
                detected_name = ''
            # Remove jika terlihat seperti ID
            elif is_likely_id_number(detected_name):
                detected_name = ''
            # Remove jika lebih dari 30% angka
            elif len(re.findall(r'\d', detected_name)) > len(detected_name) * 0.3:
                detected_name = ''
            # Remove jika mengandung keyword instansi (ini bukan nama orang!)
            elif any(kw in detected_name_lower for kw in institution_keywords_check):
                # Ini kemungkinan instansi yang salah diklasifikasi sebagai nama
                # Tukar dengan instansi jika instansi kosong atau sama
                if not detected_institution or detected_institution == detected_name:
                    detected_institution = detected_name
                detected_name = ''
            # Remove jika terlalu panjang (kemungkinan bukan nama orang, mungkin instansi)
            elif len(detected_name.split()) > 5:
                # Terlalu panjang untuk nama orang, kemungkinan ini instansi
                if not detected_institution:
                    detected_institution = detected_name
                detected_name = ''
        
        # Filter instansi: harus minimal 3 karakter, pastikan tidak mengandung nama orang di akhir
        if detected_institution:
            if len(detected_institution) < 3:
                detected_institution = ''
            elif is_likely_id_number(detected_institution) and not any(kw in detected_institution.lower() for kw in institution_keywords_check):
                detected_institution = ''
            else:
                # Bersihkan nama orang dari akhir instansi jika ada
                # Jika instansi mengandung kata-kata yang mirip nama orang di akhir
                inst_words = detected_institution.split()
                
                # Jika ada 2-4 kata terakhir yang tidak mengandung keyword instansi, mungkin itu nama orang
                if len(inst_words) >= 6:  # Instansi biasanya sudah panjang, jika ada tambahan di akhir mungkin nama orang
                    # Cek 2-4 kata terakhir
                    for i in range(2, 5):
                        if len(inst_words) >= i:
                            last_words = ' '.join(inst_words[-i:])
                            last_words_lower = last_words.lower()
                            
                            # Jika kata terakhir tidak mengandung keyword instansi, mungkin nama orang
                            if not any(kw in last_words_lower for kw in institution_keywords_check):
                                # Jika kata-kata terakhir ini adalah nama orang (dan belum ada nama), pisahkan
                                if not detected_name and 2 <= len(last_words.split()) <= 4:
                                    # Pisahkan: instansi = tanpa kata terakhir, nama = kata terakhir
                                    detected_institution = ' '.join(inst_words[:-i])
                                    detected_name = last_words
                                    break
                
                # Bersihkan whitespace berlebihan
                detected_institution = ' '.join(detected_institution.split())
        
        # Final validation: pastikan nama dan instansi berbeda
        if detected_name and detected_institution:
            if detected_name == detected_institution:
                # Jika sama, berarti salah parsing
                detected_name = ''
        
        return detected_name, detected_institution, None
            
    except ImportError:
        return None, None, "Library google-generativeai belum terinstall. Jalankan: pip install google-generativeai"
    except Exception as e:
        return None, None, f"Error menggunakan Gemini AI: {str(e)}"


@csrf_exempt
@require_http_methods(["POST"])
def ocr_process_image(request):
    """View untuk memproses gambar dengan OCR dan mendeteksi nama vs instansi"""
    try:
        # Cek apakah ada file gambar
        if 'image' not in request.FILES:
            return JsonResponse({'error': 'Tidak ada gambar yang diunggah'}, status=400)
        
        image_file = request.FILES['image']
        
        # Baca gambar
        image = Image.open(BytesIO(image_file.read()))
        
        # Convert ke RGB jika perlu
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        # OCR - gunakan pipeline teroptimasi di absensi.ocr_utils
        print("[OCR] Menggunakan pipeline OCR teroptimasi (preprocess_and_ocr)...")
        ocr_meta = {}
        try:
            # jalankan OCR dengan beberapa varian preprocessing di thread pool
            ocr_text, ocr_meta = preprocess_and_ocr(image, lang='ind+eng', psm=6, oem=1, max_workers=min(2, (os.cpu_count() or 2)))
        except Exception as e:
            print(f"[OCR] Error pada pipeline preprocess_and_ocr: {e}")
            ocr_text = ''
            ocr_meta = {}

        # Jika tidak ada hasil, coba fallback bahasa Inggris
        if not ocr_text or not ocr_text.strip():
            try:
                ocr_text = pytesseract.image_to_string(image, lang='eng', config='--oem 1 --psm 3')
                ocr_method_used = 'english_fallback'
                print("[OCR] Fallback bahasa Inggris berhasil")
            except Exception as e:
                print(f"[OCR] Fallback bahasa Inggris gagal: {e}")

        # Jika masih tidak ada hasil, kembalikan error
        if not ocr_text or not ocr_text.strip():
            return JsonResponse({
                'success': False,
                'error': 'Tidak ada teks yang terdeteksi dalam gambar setelah pipeline OCR.',
            })

        # TAHAP 1: Bersihkan dan split menjadi baris (semua informasi OCR)
        lines = [line.strip() for line in ocr_text.split('\n') if line.strip()]
        if not lines:
            return JsonResponse({'success': False, 'error': 'Tidak ada teks yang terdeteksi dalam gambar'})

        full_ocr_text = '\n'.join(lines)
        
        # TAHAP 2: Kirim ke Gemini AI untuk analisis dan sortir
        detected_name = None
        detected_institution = None
        gemini_error = None
        using_ai = False
        gemini_raw_response = None
        
        try:
            # Panggil Gemini AI untuk menganalisis dan memisahkan nama vs instansi
            gemini_name, gemini_inst, gemini_error = analyze_with_gemini(full_ocr_text)
            
            if not gemini_error and (gemini_name or gemini_inst):
                # Jika Gemini berhasil memberikan hasil, gunakan hasil Gemini (prioritas utama)
                detected_name = gemini_name
                detected_institution = gemini_inst
                using_ai = True
                
                # Log untuk debugging (opsional)
                print(f"[OCR] Gemini AI Result - Nama: {detected_name}, Instansi: {detected_institution}")
            elif gemini_error:
                print(f"[OCR] Gemini AI Error: {gemini_error}")
                
        except Exception as e:
            gemini_error = f"Error menggunakan Gemini AI: {str(e)}"
            print(f"[OCR] Gemini AI Exception: {gemini_error}")
        
        # TAHAP 3: Fallback - Jika Gemini tidak tersedia atau gagal, gunakan metode tradisional
        if not detected_name and not detected_institution:
            # Deteksi nama vs instansi menggunakan metode tradisional
            detected_name, detected_institution = detect_name_vs_institution(lines)
            
            # Jika tidak terdeteksi dengan baik, gunakan heuristik lebih pintar
            if not detected_name and not detected_institution:
                # Cari nama di baris-baris yang tersedia (skip ID)
                for line in lines:
                    if not is_likely_id_number(line) and is_likely_person_name(line):
                        detected_name = line
                        break
                
                # Cari instansi di baris-baris yang tersedia (skip ID dan nama)
                # Kumpulkan semua baris yang kemungkinan instansi
                potential_inst_lines = []
                for line in lines:
                    if line != detected_name and not is_likely_id_number(line) and not is_likely_person_name(line):
                        line_lower = line.lower()
                        # Cek apakah ini kemungkinan instansi
                        has_inst_keyword = any(kw in line_lower for kw in [
                            'pt', 'cv', 'kementerian', 'dinas', 'badan', 'lembaga',
                            'departemen', 'kantor', 'perusahaan', 'instansi', 'organisasi',
                            'pemerintah', 'pemkot', 'pemkab', 'pemprov'
                        ])
                        # Jika ada keyword instansi atau panjang (lebih dari 3 kata atau > 10 karakter)
                        if has_inst_keyword or (len(line.split()) > 3) or (len(line) > 10 and not is_likely_person_name(line)):
                            potential_inst_lines.append(line)
                
                # Gabungkan semua baris instansi yang relevan
                if potential_inst_lines:
                    detected_institution = ' '.join(potential_inst_lines)
                elif not detected_name:
                    # Jika masih belum ada instansi dan tidak ada nama, ambil baris panjang yang bukan ID
                    for line in lines:
                        if not is_likely_id_number(line) and len(line.split()) > 2:
                            detected_institution = line
                            break
        
        # TAHAP 4: Siapkan response JSON untuk auto-fill form
        # Trim ocr_meta for safe JSON serialization (only include summary info)
        try:
            variants_summary = [
                {'name': v.get('name'), 'conf_avg': v.get('conf_avg'), 'error': v.get('error')} for v in ocr_meta.get('variants', [])
            ] if isinstance(ocr_meta, dict) else []
        except Exception:
            variants_summary = []

        response_data = {
            'success': True,
            'raw_text': full_ocr_text,  # Semua informasi OCR (mentah)
            'detected_name': detected_name or '',  # Nama lengkap (hasil AI setelah sortir)
            'detected_institution': detected_institution or '',  # Instansi/perusahaan (hasil AI setelah sortir)
            'all_lines': lines,  # Semua baris OCR
            'using_ai': using_ai,  # Status apakah menggunakan AI
            'ocr_stage': 'completed',  # Status: OCR selesai
            'ai_stage': 'completed' if using_ai else 'failed',  # Status: AI selesai atau gagal
            # include a trimmed OCR metadata for debugging
            'ocr_meta': {
                'best_variant': ocr_meta.get('best_variant') if isinstance(ocr_meta, dict) else None,
                'best_conf': ocr_meta.get('best_conf') if isinstance(ocr_meta, dict) else None,
                'variant_count': ocr_meta.get('variant_count') if isinstance(ocr_meta, dict) else None,
                'variants': variants_summary,
            }
        }
        
        # Tambahkan warning jika Gemini tidak tersedia tapi masih menggunakan metode tradisional
        if gemini_error:
            response_data['warning'] = f"Gemini AI tidak tersedia ({gemini_error}), menggunakan metode deteksi tradisional"
            response_data['ai_stage'] = 'failed'
        elif using_ai:
            response_data['message'] = "Data berhasil dianalisis oleh Gemini AI dan siap untuk auto-fill"
        
        # Log final result untuk debugging
        print(f"[OCR] Final Result - Nama: {detected_name}, Instansi: {detected_institution}, Using AI: {using_ai}")
        
        return JsonResponse(response_data)
        
    except Exception as e:
        error_msg = str(e)
        # Cek apakah error karena Tesseract tidak terinstall
        if 'tesseract' in error_msg.lower() and ('not found' in error_msg.lower() or 'not installed' in error_msg.lower() or 'path' in error_msg.lower()):
            return JsonResponse({
                'success': False,
                'error': 'Tesseract OCR belum terinstall atau tidak ada di PATH. Silakan install Tesseract OCR terlebih dahulu. Lihat file INSTALL_TESSERACT.md untuk panduan instalasi.',
                'install_instructions': True
            }, status=500)
        else:
            return JsonResponse({
                'success': False,
                'error': f'Error memproses gambar: {error_msg}'
            }, status=500)


def login_view(request):
    """Halaman login untuk admin"""
    if request.user.is_authenticated:
        return redirect('admin_panel')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_staff:
                login(request, user)
                messages.success(request, 'Login berhasil!')
                return redirect('admin_panel')
            else:
                messages.error(request, 'Anda tidak memiliki akses admin.')
        else:
            messages.error(request, 'Username atau password salah.')
    
    return render(request, 'absensi/login.html')


@login_required
def logout_view(request):
    """Logout admin"""
    logout(request)
    messages.success(request, 'Anda telah logout.')
    return redirect('index')


@login_required
def admin_panel(request):
    """Halaman admin panel - menampilkan tabel daftar absensi"""
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki akses admin.')
        return redirect('index')

    # Ambil semua data absensi dengan error handling
    try:
        # Batasi hanya 10 data terbaru untuk tampilan
        absensi_list = AbsensiTamu.objects.order_by('-tanggal_waktu')[:10]
        total_absensi_count = AbsensiTamu.objects.count()
    except (OperationalError, ProgrammingError) as e:
        absensi_list = []
        total_absensi_count = 0
        messages.error(request, 'Tabel database belum dibuat! Silakan jalankan migrasi.')

    # Hitung jumlah kunjungan hari ini
    try:
        today = timezone.now().date()
        start_of_day = datetime.combine(today, datetime.min.time())
        end_of_day = datetime.combine(today, datetime.max.time())
        if is_naive(start_of_day):
            start_of_day = make_aware(start_of_day)
        if is_naive(end_of_day):
            end_of_day = make_aware(end_of_day)
        end_of_day = end_of_day + timedelta(days=1) - timedelta(seconds=1)
        absensi_hari_ini_count = AbsensiTamu.objects.filter(
            tanggal_waktu__gte=start_of_day,
            tanggal_waktu__lte=end_of_day
        ).count()
    except (OperationalError, ProgrammingError):
        absensi_hari_ini_count = 0

    # Form untuk filter tanggal
    date_range_form = DateRangeForm()

    # Convert messages to serializable format for JSON
    django_messages = []
    for message in messages.get_messages(request):
        django_messages.append({
            'message': str(message),
            'tags': message.tags
        })

    return render(request, 'absensi/admin_panel.html', {
        'absensi_list': absensi_list,
        'total_absensi_count': total_absensi_count,
        'date_range_form': date_range_form,
        'django_messages': django_messages,
        'absensi_hari_ini_count': absensi_hari_ini_count
    })


@login_required
def admin_profile_view(request):
    """Halaman untuk mengubah profil admin (username/password) dan menambah admin baru"""
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki akses yang diperlukan.')
        return redirect('index')

    if request.method == 'POST':
        # Check which form is submitted
        if 'update_profile' in request.POST:
            form = AdminProfileForm(request.POST, instance=request.user)
            if form.is_valid():
                user = form.save()
                new_password = form.cleaned_data.get('new_password1')
                if new_password:
                    update_session_auth_hash(request, user)
                    messages.success(request, 'Password Anda berhasil diubah.')
                
                messages.success(request, 'Profil Anda berhasil diperbarui.')
                return redirect('admin_profile')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        field_label = form.fields[field].label or field.replace('_', ' ').title()
                        messages.error(request, f"{field_label}: {error}")
        
        elif 'add_admin' in request.POST:
            add_admin_form = AddAdminForm(request.POST)
            if add_admin_form.is_valid():
                add_admin_form.save()
                messages.success(request, f"Admin '{add_admin_form.cleaned_data['username']}' berhasil ditambahkan.")
                return redirect('admin_profile')
            else:
                # Pass the form with errors back to the template
                form = AdminProfileForm(instance=request.user)
                return render(request, 'absensi/admin_profile.html', {
                    'form': form,
                    'add_admin_form': add_admin_form
                })

    form = AdminProfileForm(instance=request.user)
    add_admin_form = AddAdminForm()

    return render(request, 'absensi/admin_profile.html', {
        'form': form,
        'add_admin_form': add_admin_form
    })




@login_required
def download_pdf_rekap_hari_ini(request):
    """Generate dan download PDF rekapitulasi absensi berdasarkan tanggal range"""
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki akses admin.')
        return redirect('index')
    
    # Ambil parameter tanggal dari GET request atau default hari ini
    tanggal_mulai_str = request.GET.get('tanggal_mulai', '')
    tanggal_akhir_str = request.GET.get('tanggal_akhir', '')
    
    try:
        if tanggal_mulai_str and tanggal_akhir_str:
            # Jika ada parameter tanggal, gunakan range tersebut
            tanggal_mulai = datetime.strptime(tanggal_mulai_str, '%Y-%m-%d').date()
            tanggal_akhir = datetime.strptime(tanggal_akhir_str, '%Y-%m-%d').date()
            
            # Validasi tanggal
            if tanggal_mulai > tanggal_akhir:
                messages.error(request, 'Tanggal mulai tidak boleh lebih besar dari tanggal akhir.')
                return redirect('admin_panel')
        else:
            # Default: hari ini
            tanggal_mulai = timezone.now().date()
            tanggal_akhir = timezone.now().date()
    except ValueError:
        messages.error(request, 'Format tanggal tidak valid.')
        return redirect('admin_panel')
    
    # Untuk MySQL dan database lain, gunakan range datetime yang jelas
    # Mulai dari jam 00:00:00 tanggal mulai sampai 23:59:59 tanggal akhir
    start_of_range = datetime.combine(tanggal_mulai, datetime.min.time())
    end_of_range = datetime.combine(tanggal_akhir, datetime.max.time())
    
    # Convert ke timezone aware jika perlu
    if is_naive(start_of_range):
        start_of_range = make_aware(start_of_range)
    if is_naive(end_of_range):
        end_of_range = make_aware(end_of_range)
    
    # End of range harus sampai akhir hari (23:59:59), jadi tambahkan 1 detik sebelum hari berikutnya
    end_of_range = end_of_range + timedelta(days=1) - timedelta(seconds=1)
    
    try:
        # Query data berdasarkan range tanggal
        absensi_hari_ini = AbsensiTamu.objects.filter(
            tanggal_waktu__gte=start_of_range,
            tanggal_waktu__lte=end_of_range
        ).order_by('tanggal_waktu')
    except (OperationalError, ProgrammingError):
        messages.error(request, 'Error mengambil data dari database.')
        return redirect('admin_panel')
    
    # Buat response PDF
    response = HttpResponse(content_type='application/pdf')
    if tanggal_mulai == tanggal_akhir:
        filename = f"Rekap_Absensi_{tanggal_mulai.strftime('%Y-%m-%d')}.pdf"
    else:
        filename = f"Rekap_Absensi_{tanggal_mulai.strftime('%Y-%m-%d')}_to_{tanggal_akhir.strftime('%Y-%m-%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Buat dokumen PDF
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    # Container untuk elemen-elemen PDF
    elements = []
    
    # Style untuk teks
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#0066cc'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    normal_style = styles['Normal']
    normal_style.fontSize = 10
    
    # Header - Judul
    title = Paragraph("REKAPITULASI DATA ABSENSI TAMU", title_style)
    elements.append(title)
    
    # Informasi tanggal
    if tanggal_mulai == tanggal_akhir:
        tanggal_text = f"Tanggal: {tanggal_mulai.strftime('%d %B %Y')}"
    else:
        tanggal_text = f"Periode: {tanggal_mulai.strftime('%d %B %Y')} - {tanggal_akhir.strftime('%d %B %Y')}"
    tanggal_para = Paragraph(tanggal_text, normal_style)
    elements.append(tanggal_para)
    
    elements.append(Spacer(1, 0.2*inch))
    
    # Informasi statistik
    total_absensi = absensi_hari_ini.count()
    if tanggal_mulai == tanggal_akhir:
        stat_text = f"Total Absensi: <b>{total_absensi}</b> orang"
    else:
        stat_text = f"Total Absensi: <b>{total_absensi}</b> orang"
    stat_para = Paragraph(stat_text, normal_style)
    elements.append(stat_para)
    elements.append(Spacer(1, 0.3*inch))
    
    # Style untuk cell tabel dengan leading yang cukup untuk multi-line
    cell_style = ParagraphStyle(
        'TableCell',
        parent=normal_style,
        fontSize=9,
        leading=11,  # Line height, cukup untuk wrap teks
        leftIndent=0,
        rightIndent=0,
        alignment=TA_LEFT,
        spaceBefore=0,
        spaceAfter=0,
        # wordWrap sudah otomatis dilakukan oleh Paragraph berdasarkan lebar kolom
    )
    
    header_cell_style = ParagraphStyle(
        'TableHeader',
        parent=normal_style,
        fontSize=10,
        leading=12,
        textColor=colors.white,
        alignment=TA_LEFT,
        fontName='Helvetica-Bold'
    )
    
    # Jika ada data, buat tabel
    if absensi_hari_ini.exists():
        # Header tabel - menggunakan Paragraph untuk konsistensi
        header_data = [
            Paragraph('No', header_cell_style),
            Paragraph('Nama Lengkap', header_cell_style),
            Paragraph('Jabatan', header_cell_style),
            Paragraph('Instansi/Asal', header_cell_style),
            Paragraph('Alamat', header_cell_style),
            Paragraph('Nomor Telepon', header_cell_style),
            Paragraph('Maksud & Tujuan', header_cell_style),
            Paragraph('Tanggal & Waktu', header_cell_style)
        ]
        data = [header_data]
        
        # Data absensi - menggunakan Paragraph agar teks bisa wrap
        for idx, absensi in enumerate(absensi_hari_ini, start=1):
            # Escape HTML characters untuk keamanan dan formatting yang benar
            nama = escape(str(absensi.nama))
            jabatan = escape(str(absensi.jabatan)) if absensi.jabatan else '-'
            instansi = escape(str(absensi.instansi)) if absensi.instansi else '-'
            alamat = escape(str(absensi.alamat)) if absensi.alamat else '-'
            nomor_telepon = escape(str(absensi.nomor_telepon)) if absensi.nomor_telepon else '-'
            keperluan = escape(str(absensi.keperluan))  # Tidak dipotong, biarkan wrap
            waktu = absensi.tanggal_waktu.strftime('%d/%m/%Y %H:%M')
            
            # Gunakan Paragraph untuk setiap cell agar bisa auto-wrap
            # Paragraph akan secara otomatis wrap teks panjang ke baris baru
            row = [
                Paragraph(str(idx), cell_style),
                Paragraph(nama, cell_style),
                Paragraph(jabatan, cell_style),
                Paragraph(instansi, cell_style),
                Paragraph(alamat, cell_style),
                Paragraph(nomor_telepon, cell_style),
                Paragraph(keperluan, cell_style),
                Paragraph(waktu, cell_style)
            ]
            data.append(row)
        
        # Buat tabel dengan lebar kolom yang lebih proporsional
        # Total width = ~7.5 inch (A4 width - margins)
        # Kolom Keperluan dibuat lebih lebar karena biasanya teksnya paling panjang
        table = Table(
            data,
            colWidths=[0.4*inch, 1.0*inch, 0.8*inch, 1.0*inch, 1.0*inch, 0.8*inch, 1.5*inch, 1.0*inch]
        )
        
        # Style tabel
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),  # Top align untuk memungkinkan wrapping
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        
        elements.append(table)
    else:
        # Jika tidak ada data
        no_data_para = Paragraph(
            "<i>Tidak ada data absensi untuk hari ini.</i>",
            normal_style
        )
        elements.append(no_data_para)
    
    elements.append(Spacer(1, 0.5*inch))
    
    # Footer - info generate
    generate_info = f"Dicetak pada: {timezone.now().strftime('%d %B %Y, %H:%M:%S')}"
    footer_style = ParagraphStyle(
        'Footer',
        parent=normal_style,
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    footer_para = Paragraph(generate_info, footer_style)
    elements.append(footer_para)
    
    # Build PDF
    doc.build(elements)
    
    return response


@login_required
def download_excel_rekap_hari_ini(request):
    """Generate dan download Excel rekapitulasi absensi berdasarkan tanggal range"""
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki akses admin.')
        return redirect('index')
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Library openpyxl belum terinstall. Jalankan: pip install openpyxl')
        return redirect('admin_panel')
    
    # Ambil parameter tanggal dari GET request atau default hari ini
    tanggal_mulai_str = request.GET.get('tanggal_mulai', '')
    tanggal_akhir_str = request.GET.get('tanggal_akhir', '')
    
    try:
        if tanggal_mulai_str and tanggal_akhir_str:
            # Jika ada parameter tanggal, gunakan range tersebut
            tanggal_mulai = datetime.strptime(tanggal_mulai_str, '%Y-%m-%d').date()
            tanggal_akhir = datetime.strptime(tanggal_akhir_str, '%Y-%m-%d').date()
            
            # Validasi tanggal
            if tanggal_mulai > tanggal_akhir:
                messages.error(request, 'Tanggal mulai tidak boleh lebih besar dari tanggal akhir.')
                return redirect('admin_panel')
        else:
            # Default: hari ini
            tanggal_mulai = timezone.now().date()
            tanggal_akhir = timezone.now().date()
    except ValueError:
        messages.error(request, 'Format tanggal tidak valid.')
        return redirect('admin_panel')
    
    # Untuk MySQL dan database lain, gunakan range datetime yang jelas
    # Mulai dari jam 00:00:00 tanggal mulai sampai 23:59:59 tanggal akhir
    start_of_range = datetime.combine(tanggal_mulai, datetime.min.time())
    end_of_range = datetime.combine(tanggal_akhir, datetime.max.time())
    
    # Convert ke timezone aware jika perlu
    if is_naive(start_of_range):
        start_of_range = make_aware(start_of_range)
    if is_naive(end_of_range):
        end_of_range = make_aware(end_of_range)
    
    # End of range harus sampai akhir hari (23:59:59), jadi tambahkan 1 detik sebelum hari berikutnya
    end_of_range = end_of_range + timedelta(days=1) - timedelta(seconds=1)
    
    try:
        # Query data berdasarkan range tanggal
        absensi_hari_ini = AbsensiTamu.objects.filter(
            tanggal_waktu__gte=start_of_range,
            tanggal_waktu__lte=end_of_range
        ).order_by('tanggal_waktu')
    except (OperationalError, ProgrammingError):
        messages.error(request, 'Error mengambil data dari database.')
        return redirect('admin_panel')
    
    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    if tanggal_mulai == tanggal_akhir:
        ws.title = f"Rekap Absensi {tanggal_mulai.strftime('%Y-%m-%d')}"
    else:
        ws.title = f"Rekap Absensi {tanggal_mulai.strftime('%Y-%m-%d')} to {tanggal_akhir.strftime('%Y-%m-%d')}"
    
    # Style untuk header
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Style untuk data cells
    data_font = Font(name="Arial", size=10)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header - Judul dan informasi
    ws.merge_cells('A1:H1')
    ws['A1'] = "REKAPITULASI DATA ABSENSI TAMU"
    ws['A1'].font = Font(name="Arial", size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    
    ws.merge_cells('A2:H2')
    if tanggal_mulai == tanggal_akhir:
        ws['A2'] = f"Tanggal: {tanggal_mulai.strftime('%d %B %Y')}"
    else:
        ws['A2'] = f"Periode: {tanggal_mulai.strftime('%d %B %Y')} - {tanggal_akhir.strftime('%d %B %Y')}"
    ws['A2'].font = Font(name="Arial", size=10)
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
    
    ws.merge_cells('A3:H3')
    ws['A3'] = f"Total Absensi: {absensi_hari_ini.count()} orang"
    ws['A3'].font = Font(name="Arial", size=10, bold=True)
    ws['A3'].alignment = Alignment(horizontal="left", vertical="center")
    
    # Header tabel (mulai dari baris 5)
    headers = ['No', 'Nama Lengkap', 'Jabatan', 'Instansi/Asal', 'Alamat', 'Nomor Telepon', 'Maksud & Tujuan', 'Tanggal & Waktu']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data absensi
    if absensi_hari_ini.exists():
        for idx, absensi in enumerate(absensi_hari_ini, start=1):
            row_num = 5 + idx  # Mulai dari baris 6
            
            # No
            ws.cell(row=row_num, column=1, value=idx)
            
            # Nama Lengkap
            ws.cell(row=row_num, column=2, value=absensi.nama)
            
            # Jabatan
            ws.cell(row=row_num, column=3, value=absensi.jabatan or '-')
            
            # Instansi/Asal
            ws.cell(row=row_num, column=4, value=absensi.instansi if absensi.instansi else '-')
            
            # Alamat
            ws.cell(row=row_num, column=5, value=absensi.alamat if absensi.alamat else '-')
            
            # Nomor Telepon
            ws.cell(row=row_num, column=6, value=absensi.nomor_telepon if absensi.nomor_telepon else '-')
            
            # Keperluan
            ws.cell(row=row_num, column=7, value=absensi.keperluan)
            
            # Waktu
            ws.cell(row=row_num, column=8, value=absensi.tanggal_waktu.strftime('%d/%m/%Y %H:%M'))
            
            # Apply style untuk semua cell di baris ini
            for col_num in range(1, 9):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Alternating row colors
                if idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    else:
        # Jika tidak ada data
        ws.merge_cells('A6:H6')
        ws['A6'] = "Tidak ada data absensi untuk hari ini."
        ws['A6'].font = Font(name="Arial", size=10, italic=True)
        ws['A6'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Footer - info generate
    last_row = 5 + absensi_hari_ini.count() + 2 if absensi_hari_ini.exists() else 7
    ws.merge_cells(f'A{last_row}:H{last_row}')
    ws[f'A{last_row}'] = f"Dicetak pada: {timezone.now().strftime('%d %B %Y, %H:%M:%S')}"
    ws[f'A{last_row}'].font = Font(name="Arial", size=8, italic=True, color="808080")
    ws[f'A{last_row}'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    column_widths = [8, 22, 18, 24, 28, 16, 40, 18]  # Width in characters
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Set row heights
    ws.row_dimensions[1].height = 30  # Title row
    ws.row_dimensions[5].height = 25  # Header row
    
    # Buat response Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    if tanggal_mulai == tanggal_akhir:
        filename = f"Rekap_Absensi_{tanggal_mulai.strftime('%Y-%m-%d')}.xlsx"
    else:
        filename = f"Rekap_Absensi_{tanggal_mulai.strftime('%Y-%m-%d')}_to_{tanggal_akhir.strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Simpan workbook ke response
    wb.save(response)
    
    return response

